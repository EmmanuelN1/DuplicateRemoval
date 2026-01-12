# import argparse
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
# import os


# def highlight_and_sort_duplicates(csv_path):
#     # Read CSV
#     df = pd.read_csv(csv_path)

#     # Normalize first column to lowercase
#     first_col = df.columns[0]
#     df[first_col] = df[first_col].astype(str).str.lower()

#     # Identify duplicates (all occurrences)
#     duplicate_mask = df[first_col].duplicated(keep=False)

#     # Count unique duplicate values
#     unique_duplicate_count = df.loc[duplicate_mask, first_col].nunique()

#     # Move duplicates to top
#     df_sorted = pd.concat([
#         df[duplicate_mask],
#         df[~duplicate_mask]
#     ], ignore_index=True)

#     # Track duplicate row indices AFTER sorting
#     duplicate_indices = df_sorted.index[df_sorted[first_col].duplicated(keep=False)].tolist()

#     # Save to Excel
#     output_path = os.path.splitext(csv_path)[0] + "_duplicates.xlsx"
#     df_sorted.to_excel(output_path, index=False)

#     # Apply red background to duplicate rows
#     wb = load_workbook(output_path)
#     ws = wb.active

#     red_fill = PatternFill(
#         start_color="FFFF0000",
#         end_color="FFFF0000",
#         fill_type="solid"
#     )

#     for idx in duplicate_indices:
#         excel_row = idx + 2  # 1-based + header
#         for cell in ws[excel_row]:
#             cell.fill = red_fill

#     wb.save(output_path)

#     # Terminal output
#     print(f"Total UNIQUE duplicate values in first column: {unique_duplicate_count}")
#     print("Row indices of duplicate rows after sorting (0-based):")
#     print(duplicate_indices)
#     print(f"\nExcel file saved to:\n{output_path}")


# def main():
#     parser = argparse.ArgumentParser(
#         description="Move duplicates to top and highlight them in red."
#     )
#     parser.add_argument("csv_path", help="Path to the CSV file")

#     args = parser.parse_args()
#     highlight_and_sort_duplicates(args.csv_path)


# if __name__ == "__main__":
#     main()



import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


def keep_most_detailed_whitespace_duplicate(csv_path):
    df = pd.read_csv(csv_path)

    first_col = df.columns[0]

    # Preserve original value for detail comparison
    df["_original"] = df[first_col].astype(str)

    # Normalized version ONLY for duplicate detection
    df["_normalized"] = (
        df[first_col]
        .astype(str)
        .str.lower()
        .str.strip()
    )

    # Length of original string (includes whitespace)
    df["_length"] = df["_original"].str.len()

    # Identify duplicates using normalized column
    duplicate_mask = df["_normalized"].duplicated(keep=False)

    # From duplicates, keep the row with the LONGEST original string
    best_duplicates = (
        df[duplicate_mask]
        .sort_values("_length", ascending=False)
        .drop_duplicates(subset="_normalized", keep="first")
    )

    # Non-duplicate rows
    non_duplicates = df[~duplicate_mask]

    # Combine and restore original order
    final_df = (
        pd.concat([non_duplicates, best_duplicates])
        .sort_index()
        .drop(columns=["_original", "_normalized", "_length"])
    )

    # Rows to highlight
    highlight_indices = best_duplicates.index.tolist()

    # Save to Excel
    output_path = os.path.splitext(csv_path)[0] + "_cleaned.xlsx"
    final_df.to_excel(output_path, index=False)

    # Highlight kept duplicate rows
    wb = load_workbook(output_path)
    ws = wb.active

    fill = PatternFill(
        start_color="FFFFFF00",  # Yellow
        end_color="FFFFFF00",
        fill_type="solid"
    )

    for idx in highlight_indices:
        excel_row = final_df.index.get_loc(idx) + 2
        for cell in ws[excel_row]:
            cell.fill = fill

    wb.save(output_path)

    # Terminal output (no index printing)
    print(f"Total UNIQUE duplicates resolved: {len(best_duplicates)}")
    print(f"\nExcel file saved to:\n{output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Keep most detailed duplicate row (based on whitespace length)."
    )
    parser.add_argument("csv_path", help="Path to the CSV file")

    args = parser.parse_args()
    keep_most_detailed_whitespace_duplicate(args.csv_path)


if __name__ == "__main__":
    main()
